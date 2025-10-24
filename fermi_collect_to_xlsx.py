#!/usr/bin/env python3
# fermi_collect_to_xlsx.py
"""
Collect raw model responses for self-consistency annotation.

What it does
------------
- Loads CSV with columns: "Questions" and "Nearest Power of ten".
- For each question, and for each selected model (base / reasoning / both), it samples the model
  N times at a single temperature T and saves *every raw response* to an XLSX file for human annotation.
- Also writes a CSV mirror for convenience.

nohup python fermi_collect_to_xlsx.py \
  --csv_path LLN_Dataset.csv \
  --out_dir results_sc \
  --temperature 0.7 \
  --num_samples 10 \
  --models_to_run both \
  --cache_dir /home/asameull/FERMI/pm \
  --max_new_tokens 4096 >> collection.out 2>&1 & disown

CUDA_VISIBLE_DEVICES=2,3 nohup python fermi_collect_to_xlsx.py \
  --csv_path LLN_Dataset.csv \
  --out_dir results_sc/first100 \
  --start_row 0 --end_row 100 \
  --temperature 0.7 --num_samples 10 --models_to_run both \
  --cache_dir /home/asameull/FERMI/pm --max_new_tokens 4096 \
  >> first1002nd.out 2>&1 & disown

# part 1 on GPU 2
CUDA_VISIBLE_DEVICES=2 nohup python fermi_collect_to_xlsx.py \
  --csv_path LLN_Dataset.csv \
  --out_dir results_sc/second100 \
  --start_row 100 --end_row 200 \
  --temperature 0.7 --num_samples 10 --models_to_run both \
  --cache_dir /home/asameull/FERMI/pm --max_new_tokens 4096 \
  >> second100.out 2>&1 & disown

# part 2 on GPU 3
CUDA_VISIBLE_DEVICES=3 nohup python fermi_collect_to_xlsx.py \
  --csv_path LLN_Dataset.csv \
  --out_dir results_sc/third150 \
  --start_row 200 --end_row 353 \
  --temperature 0.7 --num_samples 10 --models_to_run both \
  --cache_dir /home/asameull/FERMI/pm --max_new_tokens 4096 \
  >> third150.out 2>&1 & disown

Outputs (in --out_dir)
----------------------
- responses.csv
- responses.xlsx
  - Sheet "responses": one row per (question, model_kind, sample_id)
  - Sheet "metadata": key/value run metadata
  - Sheet "codebook": column descriptions

Recommended workflow
--------------------
1) Run this script to produce responses.xlsx
2) Run the Streamlit annotator:  streamlit run annotate_fermi_streamlit.py
3) After labeling the "k_human" column, run: python fermi_finish_from_xlsx.py --xlsx results_sc/responses.xlsx

"""
import argparse
import json
import logging
import os
import sys
# os.environ["CUDA_VISIBLE_DEVICES"] = "1,2,3"

import re
from datetime import datetime
from typing import Optional, Tuple, List, Dict
from tqdm.auto import tqdm

import pandas as pd
import torch
from transformers import AutoModelForCausalLM, AutoTokenizer
try:
    # Prefer openpyxl's official regex
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
except Exception:
    # Fallback if import path changes
    ILLEGAL_CHARACTERS_RE = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")

EXCEL_MAX_CELL_CHARS = 32767  # Excel cell limit

def _sanitize_excel_str(x):
    """Remove Excel-illegal control chars and trim to Excel cell limit."""
    if x is None or isinstance(x, (int, float, bool)):
        return x
    s = str(x)
    s = ILLEGAL_CHARACTERS_RE.sub("", s)
    if len(s) > EXCEL_MAX_CELL_CHARS:
        s = s[:EXCEL_MAX_CELL_CHARS]
    return s

def _sanitize_df_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for c in out.columns:
        if out[c].dtype == object:
            out[c] = out[c].map(_sanitize_excel_str)
    return out

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S"
)
logger = logging.getLogger("fermi_collect")

def ensure_dir(path: str):
    os.makedirs(path, exist_ok=True)

def extract_power_auto(text: str) -> Optional[int]:
    """Heuristic auto-extraction of exponent k from raw text. Can be overridden by human labels later."""
    if not text:
        return None
    t = text.strip()

    m = re.search(r'\\b1(?:\\.\\d+)?e(-?\\d+)\\b', t, flags=re.IGNORECASE)
    if m:
        try:
            return int(m.group(1))
        except Exception:
            pass

    for pat in [r'10\\s*\\^\\s*(-?\\d+)', r'10\\s*\\*\\*\\s*(-?\\d+)']:
        m = re.search(pat, t, flags=re.IGNORECASE)
        if m:
            try:
                return int(m.group(1))
            except Exception:
                pass

    for pat in [
        r'(?:power\\s*of\\s*ten|nearest\\s*power\\s*of\\s*ten)\\s*[:\\-]\\s*(-?\\d+)',
        r'answer\\s*[:\\-]\\s*(-?\\d+)',
        r'\\bexponent\\s*[:\\-]\\s*(-?\\d+)',
        r'^\\s*(-?\\d+)\\s*$'
    ]:
        m = re.search(pat, t, flags=re.IGNORECASE | re.MULTILINE)
        if m:
            try:
                return int(m.group(1))
            except Exception:
                pass
    return None

def load_model(model_id: str, cache_dir: Optional[str]) -> Tuple[AutoModelForCausalLM, AutoTokenizer]:
    tok = AutoTokenizer.from_pretrained(model_id, cache_dir=cache_dir, trust_remote_code=True)
    if tok.pad_token_id is None and tok.eos_token_id is not None:
        tok.pad_token_id = tok.eos_token_id
    model = AutoModelForCausalLM.from_pretrained(
        model_id,
        cache_dir=cache_dir,
        torch_dtype=torch.float16 if torch.cuda.is_available() else None,
        device_map="auto" if torch.cuda.is_available() else None,
        trust_remote_code=True
    )
    return model, tok

def format_prompt(question: str, cot: bool) -> str:
    base = (
        "Provide the nearest power of ten for the quantity below.\\n"
        "Return ONLY the exponent k as an integer if possible (e.g., -3 for 10^-3, 6 for 10^6).\\n\\n"
        f"Question: {question}\\n\\nAnswer:"
    )
    if cot:
        return base + "\\n\\nLet me think step by step:"
    return base

@torch.inference_mode()
def generate_once(model, tokenizer, prompt: str, temperature: float, max_new_tokens: int = 512) -> str:
    inputs = tokenizer(prompt, return_tensors="pt")
    if hasattr(model, "device"):
        inputs = {k: v.to(model.device) for k, v in inputs.items()}
    outputs = model.generate(
        **inputs,
        max_new_tokens=max_new_tokens,
        do_sample=True,
        temperature=temperature,
        top_p=0.95,
        pad_token_id=tokenizer.pad_token_id or tokenizer.eos_token_id,
        eos_token_id=tokenizer.eos_token_id,
    )
    txt = tokenizer.decode(outputs[0], skip_special_tokens=True)
    if txt.startswith(prompt):
        txt = txt[len(prompt):].strip()
    return txt

def write_xlsx(out_path: str, df: pd.DataFrame, metadata: Dict[str, str], codebook: pd.DataFrame):
    from openpyxl.utils import get_column_letter
    df = _sanitize_df_for_excel(df)
    meta_df = pd.DataFrame({"key": list(metadata.keys()), "value": list(metadata.values())})
    meta_df = _sanitize_df_for_excel(meta_df)
    codebook = _sanitize_df_for_excel(codebook)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="responses")
        # meta_df = pd.DataFrame({"key": list(metadata.keys()), "value": list(metadata.values())})
        meta_df.to_excel(writer, index=False, sheet_name="metadata")
        codebook.to_excel(writer, index=False, sheet_name="codebook")

        wb = writer.book
        ws = writer.sheets["responses"]
        ws.freeze_panes = "A2"
        wide_cols = ["question", "prompt", "raw_response", "notes"]
        for col_idx, col_name in enumerate(df.columns, start=1):
            if col_name in wide_cols:
                ws.column_dimensions[get_column_letter(col_idx)].width = 80
            elif col_name in ["k_auto", "k_human", "true_k", "sample_id", "q_index"]:
                ws.column_dimensions[get_column_letter(col_idx)].width = 12
            else:
                ws.column_dimensions[get_column_letter(col_idx)].width = 20

def main():
    ap = argparse.ArgumentParser(description="Collect raw model responses to XLSX for manual annotation.")
    ap.add_argument("--csv_path", type=str, default="LLN_Dataset.csv")
    ap.add_argument("--out_dir", type=str, default="results_sc")
    ap.add_argument("--temperature", type=float, default=0.7)
    ap.add_argument("--num_samples", type=int, default=10)
    ap.add_argument("--base_model", type=str, default="Qwen/Qwen-7B")
    ap.add_argument("--reasoning_model", type=str, default="deepseek-ai/DeepSeek-R1-Distill-Qwen-7B")
    ap.add_argument("--models_to_run", choices=["base", "reasoning", "both"], default="both")
    ap.add_argument("--max_new_tokens", type=int, default=4096)
    ap.add_argument("--cache_dir", type=str, default="./pm")
    ap.add_argument("--limit", type=int, default=None)
    ap.add_argument("--start_row", type=int, default=None, help="0-based inclusive start row")
    ap.add_argument("--end_row",   type=int, default=None, help="exclusive end row")
    args = ap.parse_args()

    ensure_dir(args.out_dir)
    df = pd.read_csv(args.csv_path, encoding="windows-1252")

    device = "cuda" if torch.cuda.is_available() else "cpu"
    logger.info(f"Device: {device} (GPUs: {torch.cuda.device_count() if device == 'cuda' else 0})")

    models = {}
    if args.models_to_run in ("base", "both"):
      logger.info("Loading base model...")
      models["base"] = load_model(args.base_model, cache_dir=args.cache_dir), False
    if args.models_to_run in ("reasoning", "both"):
      logger.info("Loading reasoning model...")
      models["reasoning"] = load_model(args.reasoning_model, cache_dir=args.cache_dir), True

    rows = []
    # it = df.iterrows()
    # if args.limit is not None:
    #     it = list(it)[: args.limit]
    if args.limit is not None:
        df_iter = df.head(args.limit).iterrows()
        n_rows = min(len(df), args.limit)
    else:
        df_iter = df.iterrows()
        n_rows = len(df)
        if args.start_row is not None or args.end_row is not None:
            start = 0 if args.start_row is None else max(0, args.start_row)
            end   = len(df) if args.end_row is None else min(len(df), args.end_row)
            df_part = df.iloc[start:end]
            n_rows = len(df_part)
            df_iter = df_part.iterrows()

    row_id = 0
    for idx, r in tqdm(
                            df_iter,
                            total=n_rows,
                            desc="Questions"
                            # disable=not sys.stdout.isatty(),  # don’t spam logs under nohup
                        ):
        q = str(r.get("Questions", "")).strip()
        if not q:
            continue
        true_k = r.get("Nearest Power of ten", None)
        try:
            true_k = int(true_k)
        except Exception:
            true_k = None

        for kind, (pack, use_cot) in models.items():
            model, tok = pack
            prompt = format_prompt(q, cot=use_cot)

            for s in range(args.num_samples):
                raw = generate_once(model, tok, prompt, temperature=args.temperature, max_new_tokens=args.max_new_tokens)
                raw = _sanitize_excel_str(raw)
                k_auto = extract_power_auto(raw)
                rows.append({
                    "row_id": row_id,
                    "q_index": idx,
                    "model_kind": kind,
                    "sample_id": s,
                    "temperature": args.temperature,
                    "question": q,
                    "true_k": true_k,
                    "prompt": prompt,
                    "raw_response": raw,
                    "k_auto": k_auto,
                    "k_human": None,
                    "notes": ""
                })
                row_id += 1

        if len(rows) % 200 == 0:
            tmp = pd.DataFrame(rows)
            tmp.to_csv(os.path.join(args.out_dir, "responses.csv"), index=False)

    out_csv = os.path.join(args.out_dir, "responses.csv")
    out_xlsx = os.path.join(args.out_dir, "responses.xlsx")

    df_out = pd.DataFrame(rows)
    df_out_sanitized = _sanitize_df_for_excel(df_out)
    df_out_sanitized.to_csv(out_csv, index=False)

    codebook = pd.DataFrame([
        {"column": "row_id", "description": "Unique row identifier for this file."},
        {"column": "q_index", "description": "Original row index of question from CSV."},
        {"column": "model_kind", "description": "Model used: base or reasoning."},
        {"column": "sample_id", "description": "Self-consistency sample number (0..N-1)."},
        {"column": "temperature", "description": "Sampling temperature used for this response."},
        {"column": "question", "description": "Question text."},
        {"column": "true_k", "description": "Ground-truth nearest power-of-ten exponent (if available)."},
        {"column": "prompt", "description": "Exact prompt given to the model."},
        {"column": "raw_response", "description": "FULL raw text returned by the model."},
        {"column": "k_auto", "description": "Auto-extracted exponent (heuristic)."},
        {"column": "k_human", "description": "Human-labeled exponent (fill this in!)."},
        {"column": "notes", "description": "Optional freeform notes."},
    ])

    metadata = {
        "run_started_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "csv_path": os.path.abspath(args.csv_path),
        "out_dir": os.path.abspath(args.out_dir),
        "temperature": str(args.temperature),
        "num_samples": str(args.num_samples),
        "base_model": args.base_model,
        "reasoning_model": args.reasoning_model,
        "models_to_run": args.models_to_run,
        "max_new_tokens": str(args.max_new_tokens),
        "cache_dir": os.path.abspath(args.cache_dir) if args.cache_dir else "",
        "limit": str(args.limit) if args.limit is not None else ""
    }

    write_xlsx(out_xlsx, df_out_sanitized, metadata, codebook)
    logger.info(f"Wrote: {out_csv}")
    logger.info(f"Wrote: {out_xlsx}")

if __name__ == "__main__":
    main()

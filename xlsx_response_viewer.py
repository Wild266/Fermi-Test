#!/usr/bin/env python3
"""
xlsx_response_viewer.py

A curses-based CLI viewer/editor for an XLSX sheet named (by default)
'responses100-200.xlsx'. It focuses on three columns:

  - "raw_response"  (long text to read/scroll)
  - "prompt"        (peek at the original question)
  - "k_human"       (fill in human label; won't overwrite existing unless you confirm)

It saves **immediately** to the XLSX file after each change to avoid losing work,
and writes only the single edited cell to avoid touching anything else.

USAGE
-----
    python xlsx_response_viewer.py
    # or specify a file and/or sheet:
    python xlsx_response_viewer.py --xlsx path/to/file.xlsx --sheet Sheet1

KEYS
----
    Up / Down or j / k      : Move to previous/next row
    PgUp / PgDn             : Scroll within the current raw_response
    w / s                   : Scroll up/down one wrapped line within raw_response
    p                       : Toggle "peek" to show the prompt (question) panel
    a                       : Add k_human **only if it's empty** (quick-add)
    e                       : Edit k_human (will confirm before overwriting existing)
    m                       : Jump to next row with missing k_human
    b                       : Jump to previous row with missing k_human
    g                       : Go to a specific index (1-based within data) or Excel row #
    r                       : Reload workbook from disk (if it's changed externally)
    ?                       : Toggle help panel
    q                       : Quit

NOTES
-----
- The program looks for a header row in row 1 and expects case-insensitive names
  "raw_response", "prompt", and "k_human".
- Writes are done via openpyxl directly to the target cell, then saved atomically
  (write to a temp file and os.replace).
"""

import argparse
import curses
import os
import sys
import textwrap
from typing import Dict, List, Tuple

try:
    from openpyxl import load_workbook
except Exception as e:
    print("This tool requires 'openpyxl'. Install it with:\n    pip install openpyxl", file=sys.stderr)
    raise

# -----------------------------
# Utility: workbook I/O helpers
# -----------------------------

def _normalize(s):
    return (str(s).strip().lower() if s is not None else "")

def _build_colmap(worksheet) -> Dict[str, int]:
    """Return mapping of normalized header name -> column index (1-based)."""
    header = {}
    for j, cell in enumerate(worksheet[1], start=1):
        key = _normalize(cell.value)
        if key:
            header[key] = j
    return header

def _collect_row_indices(worksheet) -> List[int]:
    """Return a list of Excel row numbers (data rows) starting at row 2 to max_row."""
    return list(range(2, worksheet.max_row + 1))

def safe_save_atomic(wb, path: str):
    """Save workbook atomically to avoid partial writes."""
    tmp = f"{path}.tmp"
    wb.save(tmp)
    os.replace(tmp, path)

# -----------------------------
# Text wrapping / UI helpers
# -----------------------------

def wrap_text_for_width(s: str, width: int) -> List[str]:
    """Wrap text for curses display width, preserving explicit newlines."""
    if s is None:
        return ["<EMPTY>"]
    s = str(s)
    if s == "":
        return ["<EMPTY>"]
    out: List[str] = []
    for para in s.splitlines() or [""]:
        # textwrap.wrap returns [] for empty string; preserve blank lines
        wrapped = textwrap.wrap(para, width=width) if para else [""]
        out.extend(wrapped)
    return out or ["<EMPTY>"]

def clip_lines(lines: List[str], start: int, height: int) -> List[str]:
    """Return a slice of wrapped lines starting at `start` of length up to `height`."""
    if start < 0:
        start = 0
    return lines[start:start + max(0, height)]

def center_msg(width: int, msg: str) -> str:
    if len(msg) >= width:
        return msg[:width]
    pad_left = (width - len(msg)) // 2
    return " " * pad_left + msg

# -----------------------------
# Curses UI
# -----------------------------

class ViewerState:
    def __init__(self, xlsx_path: str, sheet_name: str = None):
        self.xlsx_path = xlsx_path
        self.sheet_name = sheet_name
        self.wb = None
        self.ws = None
        self.colmap: Dict[str, int] = {}
        self.rows: List[int] = []        # Excel row numbers for data rows (2..max_row)
        self.idx: int = 0                # index into self.rows (0-based)
        self.scroll: int = 0             # vertical scroll within raw_response
        self.peek: bool = False          # whether prompt panel is shown
        self.show_help: bool = False

    # --------- Workbook load/reload ----------

    def load(self):
        if not os.path.exists(self.xlsx_path):
            raise FileNotFoundError(f"File not found: {self.xlsx_path}")

        self.wb = load_workbook(self.xlsx_path)
        self.ws = self.wb[self.sheet_name] if (self.sheet_name and self.sheet_name in self.wb.sheetnames) else self.wb.active

        self.colmap = _build_colmap(self.ws)
        required = ["raw_response", "prompt", "k_human"]
        missing = [c for c in required if c not in self.colmap]
        if missing:
            raise KeyError(f"Missing required header(s): {', '.join(missing)}")

        self.rows = _collect_row_indices(self.ws)
        if not self.rows:
            raise RuntimeError("No data rows found (expected headers in row 1, data from row 2+).")

        # Bounds check current index
        self.idx = max(0, min(self.idx, len(self.rows) - 1))
        self.scroll = 0

    def reload_from_disk(self):
        """Reload workbook fresh from disk (discard in-memory unsaved changes—not expected since we save immediately)."""
        cur_excel_row = self.rows[self.idx] if self.rows else 2
        self.load()
        # Try to restore position close to previous excel row
        if self.rows:
            try:
                self.idx = max(0, self.rows.index(cur_excel_row))
            except ValueError:
                # Find nearest row
                nearest = min(self.rows, key=lambda r: abs(r - cur_excel_row))
                self.idx = self.rows.index(nearest)
        self.scroll = 0

    # --------- Data access ----------

    def get_cell(self, excel_row: int, key_norm: str):
        col = self.colmap[key_norm]
        return self.ws.cell(row=excel_row, column=col).value

    def set_cell(self, excel_row: int, key_norm: str, value):
        col = self.colmap[key_norm]
        self.ws.cell(row=excel_row, column=col).value = value
        safe_save_atomic(self.wb, self.xlsx_path)

    # --------- Navigation helpers ----------

    def current_excel_row(self) -> int:
        return self.rows[self.idx]

    def next_index(self):
        if self.idx < len(self.rows) - 1:
            self.idx += 1
            self.scroll = 0

    def prev_index(self):
        if self.idx > 0:
            self.idx -= 1
            self.scroll = 0

    def jump_next_missing(self):
        n = len(self.rows)
        start = (self.idx + 1) % n
        for step in range(n):
            j = (start + step) % n
            r = self.rows[j]
            val = self.get_cell(r, "k_human")
            if val is None or str(val).strip() == "":
                self.idx = j
                self.scroll = 0
                return True
        return False

    def jump_prev_missing(self):
        n = len(self.rows)
        start = (self.idx - 1) % n
        for step in range(n):
            j = (start - step) % n
            r = self.rows[j]
            val = self.get_cell(r, "k_human")
            if val is None or str(val).strip() == "":
                self.idx = j
                self.scroll = 0
                return True
        return False

# -----------------------------
# Input helpers
# -----------------------------

def prompt_user(stdscr, prompt: str, prefill: str = "") -> str:
    """Simple single-line input at bottom of screen with prefill."""
    curses.echo()
    curses.curs_set(1)
    h, w = stdscr.getmaxyx()
    stdscr.attrset(curses.A_REVERSE)
    stdscr.addstr(h-1, 0, " " * (w - 1))
    msg = f"{prompt}{prefill}"
    stdscr.addstr(h-1, 0, msg[:w-1])
    stdscr.move(h-1, min(len(prompt) + len(prefill), w-2))
    stdscr.attroff(curses.A_REVERSE)
    stdscr.refresh()
    # Read input; curses returns the full line when hitting Enter
    try:
        inp = stdscr.getstr(h-1, len(prompt), w - len(prompt) - 1)
        s = inp.decode("utf-8", errors="replace")
    except Exception:
        s = ""
    curses.noecho()
    curses.curs_set(0)
    return s

def confirm_overwrite(stdscr, existing_val) -> bool:
    h, w = stdscr.getmaxyx()
    msg = f"k_human already has value '{existing_val}'. Overwrite? [y/N]"
    stdscr.attrset(curses.A_REVERSE)
    stdscr.addstr(h-1, 0, " " * (w-1))
    stdscr.addstr(h-1, 0, msg[:w-1])
    stdscr.attroff(curses.A_REVERSE)
    stdscr.refresh()
    ch = stdscr.getch()
    return ch in (ord('y'), ord('Y'))

# -----------------------------
# Rendering
# -----------------------------

def draw_help(stdscr):
    h, w = stdscr.getmaxyx()
    lines = [
        "HELP",
        "",
        "Up/Down or j/k : previous/next row",
        "PgUp/PgDn      : scroll within raw_response",
        "w/s            : scroll up/down one line within raw_response",
        "p              : toggle prompt (question) pane",
        "a              : add k_human if empty",
        "e              : edit k_human (will confirm overwriting)",
        "m / b          : next / previous missing k_human",
        "g              : go to index or Excel row",
        "r              : reload workbook from disk",
        "?              : toggle this help",
        "q              : quit",
        "",
        "Press any key to close help..."
    ]
    win_h = min(len(lines) + 2, h - 4)
    win_w = min(max(len(s) for s in lines) + 4, w - 4)
    top = (h - win_h) // 2
    left = (w - win_w) // 2
    for i in range(win_h):
        stdscr.addstr(top + i, left, " " * win_w, curses.A_REVERSE)
    for i, s in enumerate(lines):
        stdscr.addstr(top + 1 + i, left + 2, s[:win_w-4], curses.A_REVERSE)
    stdscr.refresh()
    stdscr.getch()

def render(stdscr, state: ViewerState, status_msg: str = ""):
    stdscr.clear()
    h, w = stdscr.getmaxyx()

    # Header/status line (reversed)
    excel_row = state.current_excel_row()
    rr = state.get_cell(excel_row, "raw_response")
    pr = state.get_cell(excel_row, "prompt")
    kv = state.get_cell(excel_row, "k_human")
    kv_display = "<MISSING>" if (kv is None or str(kv).strip() == "") else str(kv)

    header = f" File: {os.path.basename(state.xlsx_path)}  |  Sheet: {state.ws.title}  |  Row {state.idx+1}/{len(state.rows)} (Excel {excel_row})  |  k_human: {kv_display} "
    if len(header) < w:
        header = header + " " * (w - len(header) - 1)
    stdscr.addstr(0, 0, header[:w-1], curses.A_REVERSE)

    # Help hint / status
    hint = "[j/k] rows  [PgUp/PgDn,w/s] scroll  [p] prompt  [a] add  [e] edit  [m/b] next/prev missing  [g] goto  [?] help  [q] quit"
    stdscr.addstr(1, 0, hint[:w-1])

    # Optional status message
    if status_msg:
        stdscr.addstr(2, 0, status_msg[:w-1], curses.A_BOLD)
        start_y = 3
    else:
        start_y = 2

    # Compute layout
    prompt_panel_height = max(3, h // 4) if state.peek else 0
    content_height = max(1, h - start_y - prompt_panel_height - 1)  # reserve 1 line for footer
    content_width = max(10, w - 2)

    # Render raw_response in the content area (bounded by scroll)
    wrapped_rr = wrap_text_for_width(rr, width=content_width)
    max_scroll = max(0, len(wrapped_rr) - content_height)

    # Draw border-ish edges with a minimal margin
    for i in range(content_height):
        line = wrapped_rr[state.scroll + i] if state.scroll + i < len(wrapped_rr) else ""
        stdscr.addstr(start_y + i, 1, line[:content_width])

    # If peeking, render prompt panel at bottom
    if state.peek:
        stdscr.addstr(start_y + content_height, 0, "-" * (w - 1))
        prompt_lines = wrap_text_for_width(pr, width=content_width)
        visible = clip_lines(prompt_lines, 0, prompt_panel_height - 1)
        for i, line in enumerate(visible):
            stdscr.addstr(start_y + content_height + 1 + i, 1, line[:content_width])

    # Footer line (reverse) shows scroll info
    scroll_info = f" Scroll {state.scroll}/{max_scroll} "
    footer = center_msg(w, scroll_info)
    stdscr.addstr(h-1, 0, footer[:w-1], curses.A_REVERSE)

    stdscr.refresh()
    return max_scroll

# -----------------------------
# Main event loop
# -----------------------------

def run_curses(stdscr, state: ViewerState):
    curses.curs_set(0)
    stdscr.keypad(True)
    curses.noecho()

    status_msg = ""
    max_scroll = render(stdscr, state, status_msg=status_msg)

    while True:
        ch = stdscr.getch()

        if ch in (ord('q'), ord('Q')):
            break

        elif ch in (ord('?'),):
            draw_help(stdscr)
            status_msg = ""
            max_scroll = render(stdscr, state, status_msg=status_msg)
            continue

        elif ch in (curses.KEY_RESIZE,):
            state.scroll = 0
            status_msg = ""
            max_scroll = render(stdscr, state, status_msg=status_msg)
            continue

        # Navigation rows
        elif ch in (curses.KEY_DOWN, ord('j')):
            state.next_index()
            status_msg = ""
            max_scroll = render(stdscr, state, status_msg=status_msg)
            continue

        elif ch in (curses.KEY_UP, ord('k')):
            state.prev_index()
            status_msg = ""
            max_scroll = render(stdscr, state, status_msg=status_msg)
            continue

        # Jump next/prev missing k_human
        elif ch in (ord('m'),):
            found = state.jump_next_missing()
            status_msg = "Jumped to next missing k_human" if found else "No missing k_human found."
            max_scroll = render(stdscr, state, status_msg=status_msg)
            continue

        elif ch in (ord('b'),):
            found = state.jump_prev_missing()
            status_msg = "Jumped to previous missing k_human" if found else "No missing k_human found."
            max_scroll = render(stdscr, state, status_msg=status_msg)
            continue

        # Toggle prompt peek
        elif ch in (ord('p'), ord('P')):
            state.peek = not state.peek
            status_msg = "Prompt panel ON" if state.peek else "Prompt panel OFF"
            max_scroll = render(stdscr, state, status_msg=status_msg)
            continue

        # Scroll within current raw_response
        elif ch in (curses.KEY_NPAGE,):  # PageDown
            state.scroll = min(state.scroll + max(1, (max_scroll // 2) or 1), max_scroll)
            status_msg = ""
            max_scroll = render(stdscr, state, status_msg=status_msg)
            continue

        elif ch in (curses.KEY_PPAGE,):  # PageUp
            state.scroll = max(state.scroll - max(1, (max_scroll // 2) or 1), 0)
            status_msg = ""
            max_scroll = render(stdscr, state, status_msg=status_msg)
            continue

        elif ch in (ord('w'),):
            state.scroll = max(0, state.scroll - 1)
            status_msg = ""
            max_scroll = render(stdscr, state, status_msg=status_msg)
            continue

        elif ch in (ord('s'),):
            state.scroll = min(max_scroll, state.scroll + 1)
            status_msg = ""
            max_scroll = render(stdscr, state, status_msg=status_msg)
            continue

        # Go to index / excel row
        elif ch in (ord('g'),):
            # Ask for either a "row index" (1..N) or "excel row" (#).
            s = prompt_user(stdscr, "Go to (index 1..N or Excel row #): ")
            s = s.strip()
            moved = False
            if s.isdigit():
                val = int(s)
                # Try as index first
                if 1 <= val <= len(state.rows):
                    state.idx = val - 1
                    moved = True
                else:
                    # Try as excel row
                    if val in state.rows:
                        state.idx = state.rows.index(val)
                        moved = True
            if moved:
                state.scroll = 0
                status_msg = f"Jumped to row {state.idx+1} (Excel {state.current_excel_row()})"
            else:
                status_msg = "Invalid row."
            max_scroll = render(stdscr, state, status_msg=status_msg)
            continue

        # Add k_human if empty
        elif ch in (ord('a'),):
            r = state.current_excel_row()
            existing = state.get_cell(r, "k_human")
            if existing is not None and str(existing).strip() != "":
                status_msg = f"k_human already set to '{existing}'. Use 'e' to edit."
                max_scroll = render(stdscr, state, status_msg=status_msg)
                continue
            prefill = ""  # Suggest empty
            new_val = prompt_user(stdscr, f"Set k_human for Excel row {r}: ", prefill=prefill)
            new_val = new_val.strip()
            if new_val == "":
                status_msg = "No change."
            else:
                # Write and save
                state.set_cell(r, "k_human", new_val)
                status_msg = f"Saved k_human = '{new_val}'"
            max_scroll = render(stdscr, state, status_msg=status_msg)
            continue

        # Edit k_human (with overwrite confirm when needed)
        elif ch in (ord('e'),):
            r = state.current_excel_row()
            existing = state.get_cell(r, "k_human")
            if existing is not None and str(existing).strip() != "":
                if not confirm_overwrite(stdscr, existing):
                    status_msg = "Edit cancelled."
                    max_scroll = render(stdscr, state, status_msg=status_msg)
                    continue
                prefill = str(existing)
            else:
                prefill = ""
            new_val = prompt_user(stdscr, f"Enter k_human for Excel row {r}: ", prefill=prefill)
            # If user simply hits enter with empty and there was an existing value, keep as is.
            if new_val.strip() == "" and (existing is not None and str(existing).strip() != ""):
                status_msg = "No change."
            else:
                state.set_cell(r, "k_human", new_val.strip())
                status_msg = f"Saved k_human = '{new_val.strip()}'"
            max_scroll = render(stdscr, state, status_msg=status_msg)
            continue

        # Reload workbook
        elif ch in (ord('r'),):
            try:
                state.reload_from_disk()
                status_msg = "Workbook reloaded."
            except Exception as e:
                status_msg = f"Reload failed: {e}"
            max_scroll = render(stdscr, state, status_msg=status_msg)
            continue

        # Unknown key: ignore or flash
        else:
            status_msg = ""
            max_scroll = render(stdscr, state, status_msg=status_msg)
            continue

def main():
    parser = argparse.ArgumentParser(description="CLI viewer/editor for 'raw_response'/'prompt'/'k_human' in an XLSX file.")
    parser.add_argument("--xlsx", default="responses100-200.xlsx", help="Path to the XLSX file (default: responses100-200.xlsx)")
    parser.add_argument("--sheet", default=None, help="Sheet name (default: active sheet)")
    args = parser.parse_args()

    state = ViewerState(args.xlsx, args.sheet)
    try:
        state.load()
    except Exception as e:
        print(f"Error loading workbook: {e}", file=sys.stderr)
        sys.exit(1)

    curses.wrapper(run_curses, state)

if __name__ == "__main__":
    main()
